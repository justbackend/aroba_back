import io
from datetime import datetime, date
from typing import Union, Any, TypeVar

import openpyxl
import pytz
from django.db.models import Model, QuerySet
from django.http import HttpResponse
from openpyxl.styles import Alignment, Side, PatternFill, Border, Font
from openpyxl.workbook import Workbook
from rest_framework.views import APIView

T = TypeVar('T', bound='Model')


class WriteWorkBook(object):
    DEFAULT_WIDTH = 15
    DATETIME_FORMAT = '%Y-%m-%d %H:%M:%S'
    DATE_FORMAT = '%Y-%m-%d'
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center")
    timezone = pytz.timezone('Asia/Tashkent')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(
            self,
            model,
            fields: Union[str, list],
            data=None,
            headers: list = None,
            column_widths: list = None,
    ):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.model = model
        self.data = data
        self.fields = self.set_fields(fields)
        self.headers = self.set_headers(headers)
        self.column_widths = self.set_column_widths(column_widths)

    def set_fields(self, fields):
        if fields == '__all__' and issubclass(self.model, Model):
            return [field.name for field in self.model._meta.get_fields()]
        return fields

    def set_headers(self, headers):
        if headers is None:
            return self.fields
        return headers

    def set_column_widths(self, column_widths):
        if column_widths is None:
            return (self.DEFAULT_WIDTH for i in range(len(self.fields)))
        return column_widths

    def collect_row(self, obj):
        row = []

        for field in self.fields:
            if isinstance(obj, dict) or issubclass(obj.__class__, dict):
                value = obj.get(field)
            else:
                value = getattr(obj, field, None)

            if isinstance(value, datetime):
                value = value.astimezone(self.timezone)
                value = value.strftime(self.DATETIME_FORMAT)
            elif isinstance(value, date):
                value = value.strftime(self.DATE_FORMAT)
            row.append(value)
        return row

    def set_style_to_header(self):

        for cell in self.sheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.alignment_center

    def set_thin_border(self):
        for row in self.sheet.iter_rows(min_row=2, max_col=len(self.headers), max_row=self.sheet.max_row):
            for cell in row:
                cell.alignment = self.alignment_center
                cell.border = self.thin_border

    def get_workbook_file(self, thin_border=True, set_style_header=True):
        assert self.data is not None, "There is no data"

        self.sheet.append(self.headers)
        for i, column_width in enumerate(self.column_widths, 1):
            self.sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        for obj in self.data:
            self.sheet.append(self.collect_row(obj))

        if thin_border:
            self.set_thin_border()

        if set_style_header:
            self.set_style_to_header()

        return self.workbook


class ExcelListView(APIView):
    excel_headers: list = None
    fields: list = None
    column_widths: list = None
    default_widths: int = 15
    data: Any = None
    filters: list = None
    filename: str = 'New_excel'

    def get(self, request, *args, **kwargs):
        workbook_obj = WriteWorkBook(
            model=None,
            data=self.get_data(),
            fields=self.fields,
            headers=self.excel_headers,
            column_widths=self.column_widths
        )
        workbook_obj.DEFAULT_WIDTH = self.default_widths

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        bio = io.BytesIO()

        workbook = workbook_obj.get_workbook_file()
        workbook.save(bio)

        response = response
        response['Content-Disposition'] = f'attachment; filename={self.filename}.xlsx'
        response.write(bio.getvalue())
        return response

    def get_data(self):
        assert self.data is not None, "You have to set data"

        if isinstance(self.data, QuerySet) and self.filters:
            return self.get_filtered_data(self.data)

        return self.data

    def get_filtered_data(self, data: QuerySet[T]) -> QuerySet[T]:
        query_params = self.request.query_params
        query = dict()
        for field in self.filters:
            if value := query_params.get(field):
                query[field] = value
        return data.filter(**query)
